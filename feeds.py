import csv
import openpyxl
import os
import requests
from bs4 import BeautifulSoup


class Eprice:
    @classmethod
    def eprice(cls, rev):
        abspath = os.getcwd()
        headers = None

        euro_to_czk_url = "https://www.xe.com/currencyconverter/convert/?Amount=1&From=EUR&To=CZK"

        res = requests.get(euro_to_czk_url)

        with open(f"{abspath}/html/euro_to_czk.html", "w") as file:
            file.write(res.text)

        with open(f"{abspath}/html/euro_to_czk.html") as file:
            src = file.read()

        soup = BeautifulSoup(src, "lxml")

        result = soup.find("p", class_="result__BigRate-sc-1bsijpp-1 iGrAod").text
        currency_euro_to_czk = float(result.split()[0])

        revenue = rev

        workbook = openpyxl.load_workbook("updated.xlsx")
        worksheet = workbook.active

        rix = []
        for row in worksheet:
            with open(f'{abspath}/feeds/eprice feed.csv', "r") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for r in reader:
                    x = r[0].split(";")
                    if x[0] in str(row[0].value) and len(x) == 19:
                        x[5] = str(round(((float(row[2].value) / currency_euro_to_czk)*revenue), 2))
                        x[7] = row[3].value
                        rix.append(x)
                        break
                    
        with open(f'{abspath}/feeds/eprice feed updated.csv', "a") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rix:
                writer.writerow(row)


class KaufLand:
    @classmethod
    def kaufland(cls, rev):
        abspath = os.getcwd()
        headers = None

        euro_to_czk_url = "https://www.xe.com/currencyconverter/convert/?Amount=1&From=EUR&To=CZK"

        res = requests.get(euro_to_czk_url)

        with open(f"{abspath}/html/euro_to_czk_2.html", "w") as file:
            file.write(res.text)

        with open(f"{abspath}/html/euro_to_czk_2.html") as file:
            src = file.read()

        soup = BeautifulSoup(src, "lxml")

        result = soup.find("p", class_="result__BigRate-sc-1bsijpp-1 iGrAod").text
        currency_euro_to_czk = round(float(result.split()[0]), 2)

        revenue = rev

        workbook = openpyxl.load_workbook("updated.xlsx")
        worksheet = workbook.active

        new_header = [
            "ean",
            "condition",
            "price",
            "currency",
            "id_warehouse",
            "count",
            "price_cs",
            "handling_time"
        ]

        rix = []
        rix.append(new_header) 
        for row in worksheet:
            listing = ["", "", "", "", "", "", "", ""]
            if str(row[0].value) != "sku" and len(str(row[0].value)) == 6:
                listing[0] = str(row[1].value)
                listing[1] = ''
                listing[2] = str(int(((float(row[2].value) / currency_euro_to_czk)*revenue)))
                listing[3] = currency_euro_to_czk
                listing[4] = str(row[0].value)
                listing[5] = row[3].value

                rix.append(listing)
            
        with open(f'{abspath}/feeds/kaufland-de updated.csv', "a") as f:
            writer = csv.writer(f)
            for row in rix:
                writer.writerow(row)


class Allegro:
    @classmethod
    def allegro(cls, rev):
        ##
        abspath = os.getcwd() 
        pln_to_czk_url = "https://www.xe.com/currencyconverter/convert/?Amount=1&From=PLN&To=CZK"

        res = requests.get(pln_to_czk_url)

        with open(f"{abspath}/html/pln_to_czk.html", "w") as file:
            file.write(res.text)

        with open(f"{abspath}/html/pln_to_czk.html") as file:
            src = file.read()

        soup = BeautifulSoup(src, "lxml")

        result = soup.find("p", class_="result__BigRate-sc-1bsijpp-1 iGrAod").text
        currency_pln_to_czk = round(float(result.split()[0]), 2)
        
        workbook = openpyxl.load_workbook(f"{abspath}/feeds/sddone_alegro.xlsm")
        worksheet = workbook.active
        counter = 0
        new_header = None
        revenue = rev

        for row in worksheet:
            counter += 1
            listing = []
            if counter <= 3:
                    continue
            for i in range(0, worksheet.max_row):
                if row[i].value != None:
                    if counter == 4:
                        if i > 7:
                            listing.append(row[i].value)
            if counter == 4:
                new_header = listing[:10]
                break

        workbook.close()

        workbook = openpyxl.load_workbook("updated.xlsx")
        worksheet = workbook.active
        rix = []
        rix.append(new_header) 
        for row in worksheet:
            fill = ["", "", "", "", "", "", "", "", "", ""]
            if str(row[0].value) != "sku" and len(str(row[0].value)) == 6:
                fill[0] = str(row[1].value)
                fill[3] = str(row[0].value)
                fill[4] = str(row[3].value)
                fill[5] = str(int(((float(row[2].value) / currency_pln_to_czk)*revenue)))
                rix.append(fill)

        workbook.close()

        with open(f'{abspath}/feeds/allegro updated.csv', "a") as f:
            writer = csv.writer(f)
            for row in rix:
                writer.writerow(row)
    