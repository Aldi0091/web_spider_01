import openpyxl


class Url:

    def __init__(self, path, dict):
        self.path = path
        self.dict = dict

    def collect_all_urls(self):
        out = []
        workbook = openpyxl.load_workbook(self.path)

        worksheet = workbook.active

        for row in worksheet.iter_rows(0, worksheet.max_row):
            for i in range(1, worksheet.max_column):
                if "https" in str(row[i].value):
                    out.append(row[i].value)
        return out

    def update_sheet(self):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook.active

        for key, value in self.dict.items():
            price = str(value["price without VAT"]).replace(" ", "")[:-2]
            nedostupno = 'Momentálně nedostupné - Neznáme termín dodání'
            sklad_2 = 'Skladem 2 kusy Kdy zboží dostanu?'
            sklad_3 = 'Skladem 3 kusy Kdy zboží dostanu?'
            neni = 'Není skladem'
            nezname = 'Předpokládaný termín dodání: neznámý termín dodání - Neznáme termín dodání'
            sklad_1 = '1 kus'
            do_2_dnu = 'Skladem dodání do 2 dnů Kdy zboží dostanu?'
            sklad_01 = 'Skladem 1 kus Kdy zboží dostanu?'
            on_order = 'Skladem na objednání Kdy zboží dostanu?'
            nepotvr = 'Předpokládaný termín dodání: nepotvrzený termín dodání - Neznáme termín dodání'
            sklad_4 = 'Skladem 4 kusy Kdy zboží dostanu?'
            sklad_5 = 'Skladem 5 a více kusů Kdy zboží dostanu?'
            if value["quantity"] == nedostupno or value["quantity"] == neni or value["quantity"] == nezname or value["quantity"] == nepotvr:
                quantity = 0
            elif value["quantity"] == sklad_2:
                quantity = 2
            elif value["quantity"] == sklad_3:
                quantity = 3
            elif value["quantity"] == sklad_1 or value["quantity"] == do_2_dnu or value["quantity"] == sklad_01 or value["quantity"] == on_order:
                quantity = 1
            elif value["quantity"] == sklad_4:
                quantity = 4
            elif value["quantity"] == sklad_5:
                quantity = 5
            for row in worksheet:
                if str(row[0].value) in key:
                    row[2].value = str(price) + ".00"
                    row[3].value = str(quantity)
        workbook.save("updated.xlsx")